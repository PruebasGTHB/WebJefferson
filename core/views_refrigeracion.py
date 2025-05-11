from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET
from influxdb_client import InfluxDBClient
from openpyxl import Workbook
from django.conf import settings
from datetime import datetime
import pytz
import traceback
import math

ZONA_LOCAL = pytz.timezone(settings.INFLUXDB_ZONE)

# ----------------------------
# MÉTRICAS PRINCIPALES
# ----------------------------
METRICAS = {
    "voltaje": "VOLTAJE_FASES_PROMEDIO",
    "amperaje": "CORRIENTE_PROMEDIO",
    "demanda": "POTENCIA_ACTIVA_TOTAL"
}

COMPRESORES = ["C1", "C2", "C3", "C4", "C5", "C6", "C7", "C8"]
TUNELES = ["T1", "T2", "T3", "T4", "T5", "T6", "T7", "T8", "T9", "T10"]

# ----------------------------
# FUNCIONES AUXILIARES
# ----------------------------

def limpiar_nan(valores):
    resultado = []
    for v in valores:
        if isinstance(v, (int, float)):
            if math.isnan(v):
                resultado.append(None)
            else:
                valor = max(0, round(v, 2))  # 👈 fuerza mínimo 0
                resultado.append(valor)
        else:
            resultado.append(None)
    return resultado


from pandas import DataFrame

def metricas_por_metrica(bucket, tag_key, tags, field, inicio, fin):
    from influxdb_client import InfluxDBClient
    import pandas as pd

    tag_filter = " or ".join([f'r["{tag_key}"] == "{tag}"' for tag in tags])
    query = f'''
        from(bucket: "{bucket}")
        |> range(start: time(v: "{inicio}T00:00:00Z"), stop: time(v: "{fin}T23:59:59Z"))
        |> filter(fn: (r) => r._measurement == "sensor_data" and r._field == "{field}" and ({tag_filter}))
        |> aggregateWindow(every: 1h, fn: mean, createEmpty: false)
    '''

    with InfluxDBClient(
        url=settings.INFLUXDB_URL,
        token=settings.INFLUXDB_TOKEN,
        org=settings.INFLUXDB_ORG
    ) as client:
        df = client.query_api().query_data_frame(query)

    if df.empty:
        return []

    df["_time"] = df["_time"].dt.tz_convert(ZONA_LOCAL)
    df["_value"] = pd.to_numeric(df["_value"], errors="coerce")
    df = df.dropna(subset=["_value"])

    resultados = []
    for tag in tags:
        df_tag = df[df[tag_key] == tag]
        if df_tag.empty:
            continue
        valores = df_tag["_value"].tolist()
        fechas = df_tag["_time"].dt.strftime("%Y-%m-%d %H:%M").tolist()
        valores_limpios = limpiar_nan(valores)
        valores_numericos = [v for v in valores_limpios if v is not None]
        if not valores_numericos:
            continue
        resultados.append({
            "tag": tag,
            "fechas": fechas,
            "valores": valores_limpios,
            "min": round(min(valores_numericos), 2),
            "max": round(max(valores_numericos), 2),
            "prom": round(sum(valores_numericos) / len(valores_numericos), 2)
        })

    return resultados

def graficos_por_bucket(bucket, tag_key, fields, inicio, fin):
    from influxdb_client import InfluxDBClient
    import pandas as pd

    field_filter = " or ".join([f'r._field == "{f}"' for f in fields])
    query = f'''
        from(bucket: "{bucket}")
        |> range(start: time(v: "{inicio}T00:00:00Z"), stop: time(v: "{fin}T23:59:59Z"))
        |> filter(fn: (r) => r._measurement == "sensor_data" and ({field_filter}))
        |> aggregateWindow(every: 1h, fn: mean, createEmpty: false)
        |> pivot(rowKey: ["_time"], columnKey: ["_field", "{tag_key}"], valueColumn: "_value")
    '''

    with InfluxDBClient(
        url=settings.INFLUXDB_URL,
        token=settings.INFLUXDB_TOKEN,
        org=settings.INFLUXDB_ORG
    ) as client:
        df = client.query_api().query_data_frame(query)

    if df.empty:
        return {f: {"fechas": [], "series": []} for f in fields}

    df["_time"] = df["_time"].dt.tz_convert(ZONA_LOCAL)
    fechas = df["_time"].dt.strftime("%Y-%m-%d %H:%M").tolist()

    resultado = {}
    for field in fields:
        columnas = [
            col for col in df.columns
            if (isinstance(col, tuple) and col[0] == field)
            or (isinstance(col, str) and col.startswith(field))
        ]
        series = []
        for col in columnas:
            # Extraer el tag
            if isinstance(col, tuple):
                tag = col[1]
            elif isinstance(col, str):
                tag = col.replace(f"{field}_", "")
            else:
                continue

            valores = limpiar_nan(df[col].tolist())
            if any(v is not None for v in valores):
                series.append({"nombre": tag, "valores": valores})
        resultado[field] = {"fechas": fechas, "series": series}

    return resultado

# ----------------------------
# VISTA PRINCIPAL
# ----------------------------

from django.core.cache import cache
from datetime import date
import time

@csrf_exempt
def datos_refrigeracion(request):
    try:
        t0 = time.time()

        inicio = request.GET.get("inicio", "2025-01-01")
        fin = request.GET.get("fin", "2025-01-31")
        cache_key = f"datos_refrigeracion_{inicio}_{fin}"

        cached = cache.get(cache_key)
        if cached:
            print(f"✅ Cache HIT para {inicio} a {fin} en {round((time.time() - t0)*1000)}ms")
            return JsonResponse(cached)

        # ----------------------------
        # Generar datos
        # ----------------------------
        metricas_comp = {
            k: metricas_por_metrica("Compresores", "topic", COMPRESORES, v, inicio, fin)
            for k, v in METRICAS.items()
        }
        metricas_tun = {
            k: metricas_por_metrica("Tuneles", "TAG", TUNELES, v, inicio, fin)
            for k, v in METRICAS.items()
        }

        graf_comp = graficos_por_bucket("Compresores", "topic", list(METRICAS.values()), inicio, fin)
        graf_tun = graficos_por_bucket("Tuneles", "TAG", list(METRICAS.values()), inicio, fin)

        result = {
            "compresores": metricas_comp,
            "tuneles": metricas_tun,
            "grafico_voltaje_compresores": graf_comp[METRICAS["voltaje"]],
            "grafico_amperaje_compresores": graf_comp[METRICAS["amperaje"]],
            "grafico_demanda_compresores": graf_comp[METRICAS["demanda"]],
            "grafico_voltaje_tuneles": graf_tun[METRICAS["voltaje"]],
            "grafico_amperaje_tuneles": graf_tun[METRICAS["amperaje"]],
            "grafico_demanda_tuneles": graf_tun[METRICAS["demanda"]]
        }

        # ----------------------------
        # Lógica de timeout del caché
        # ----------------------------
        fecha_fin = datetime.strptime(fin, "%Y-%m-%d").date()
        hoy = date.today()

        timeout_cache = 0 if fecha_fin < hoy else 60
        cache.set(cache_key, result, timeout=timeout_cache)

        print(f"⏱️ datos_refrigeracion para {inicio} a {fin} → {round((time.time() - t0)*1000)}ms")
        return JsonResponse(result)

    except Exception as e:
        print("❌ Error en datos_refrigeracion:", traceback.format_exc())
        return JsonResponse({"error": str(e)}, status=500)

# ----------------------------
# ÚLTIMA POTENCIA ACTUAL
# ----------------------------
@require_GET
def ultima_potencia(request):
    try:
        elementos = request.GET.getlist("tags[]")
        if not elementos:
            return JsonResponse({"error": "No se proporcionaron tags."}, status=400)

        resultados = {}
        with InfluxDBClient(
            url=settings.INFLUXDB_URL,
            token=settings.INFLUXDB_TOKEN,
            org=settings.INFLUXDB_ORG
        ) as client:
            query_api = client.query_api()
            for tag in elementos:
                bucket, tag_key = ("Compresores", "topic") if tag.startswith("C") else ("Tuneles", "TAG") if tag.startswith("T") else (None, None)
                if not bucket:
                    continue
                query = f'''
                    from(bucket: "{bucket}")
                    |> range(start: -2h)
                    |> filter(fn: (r) => r._measurement == "sensor_data" and r._field == "POTENCIA_ACTIVA_TOTAL" and r["{tag_key}"] == "{tag}")
                    |> last()
                '''
                tables = query_api.query(query)
                for table in tables:
                    for record in table.records:
                        val = record.get_value()
                        if isinstance(val, (int, float)) and not math.isnan(val):
                            resultados[tag] = round(val, 2)

        return JsonResponse(resultados)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

# ----------------------------
# EXPORTACIÓN DE EXCEL
# ----------------------------
@csrf_exempt
def descargar_datos(request):
    try:
        import pandas as pd
        inicio = request.GET.get("inicio", "2025-01-01")
        fin = request.GET.get("fin", "2025-01-31")
        zona = pytz.timezone(settings.INFLUXDB_ZONE)

        equipos = {
            "Compresores": ("topic", COMPRESORES),
            "Tuneles": ("TAG", TUNELES)
        }

        wb = Workbook()
        wb.remove(wb.active)

        with InfluxDBClient(url=settings.INFLUXDB_URL, token=settings.INFLUXDB_TOKEN, org=settings.INFLUXDB_ORG) as client:
            query_api = client.query_api()

            for bucket, (tag_key, tags) in equipos.items():
                # Consulta combinada por bucket
                field_filter = " or ".join([f'r._field == "{f}"' for f in METRICAS.values()])
                query = f'''
                    from(bucket: "{bucket}")
                    |> range(start: time(v: "{inicio}T00:00:00Z"), stop: time(v: "{fin}T23:59:59Z"))
                    |> filter(fn: (r) => r._measurement == "sensor_data" and ({field_filter}))
                    |> aggregateWindow(every: 1h, fn: mean, createEmpty: false)
                    |> pivot(rowKey: ["_time"], columnKey: ["_field", "{tag_key}"], valueColumn: "_value")
                '''

                df = query_api.query_data_frame(query)
                if df.empty:
                    continue

                df["_time"] = df["_time"].dt.tz_convert(zona)
                fechas = df["_time"].dt.strftime("%Y-%m-%d %H:%M")

                datos_ws = wb.create_sheet(title=f"{bucket} - Datos")
                resumen_ws = wb.create_sheet(title=f"{bucket} - Resumen")
                datos_ws.append(["Fecha", "Tag", "Métrica", "Valor"])
                resumen_ws.append(["Tag", "Métrica", "Mínimo", "Máximo", "Promedio"])

                for field, field_name in METRICAS.items():
                    columnas = [col for col in df.columns if isinstance(col, tuple) and col[0] == field_name]
                    for col in columnas:
                        tag = col[1]
                        valores = pd.to_numeric(df[col], errors="coerce").fillna(pd.NA)
                        if valores.isna().all():
                            continue
                        for fecha, valor in zip(fechas, valores):
                            if pd.notna(valor):
                                datos_ws.append([fecha, tag, field, round(valor, 2)])
                        valores_validos = valores.dropna().apply(lambda x: max(0, x))  # 👈 Fuerza mínimo 0
                        if not valores_validos.empty:
                            resumen_ws.append([
                                tag, field,
                                round(valores_validos.min(), 2),
                                round(valores_validos.max(), 2),
                                round(valores_validos.mean(), 2)
                            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"datos_refrigeracion_{inicio}_a_{fin}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    except Exception as e:
        print("❌ Error en descargar_datos:", traceback.format_exc())
        return HttpResponse(f"Error al generar Excel: {str(e)}", status=500)

